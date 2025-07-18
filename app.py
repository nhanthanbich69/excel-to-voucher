import streamlit as st
import pandas as pd
import os
import re
from io import BytesIO
from openpyxl import load_workbook

st.set_page_config(page_title="Sửa chứng từ tự động", layout="centered")

st.title("🧾 Sửa cột 'Số chứng từ' từ tên file Excel")

uploaded_files = st.file_uploader("📁 Tải lên nhiều file Excel", type=["xlsx"], accept_multiple_files=True)

if uploaded_files:
    for uploaded_file in uploaded_files:
        try:
            # 🎯 Lấy tháng-năm từ tên file, ví dụ: "Thu tiền 2022.09.xlsx"
            filename = uploaded_file.name
            match = re.search(r"(\d{4})\.(\d{2})", filename)
            if not match:
                st.warning(f"⚠️ Không tìm thấy tháng-năm trong tên file: {filename}")
                continue
            year, month = match.groups()
            prefix = f"PT_THUOC_"

            # 📄 Đọc file Excel với openpyxl để giữ định dạng gốc
            wb = load_workbook(uploaded_file)
            ws = wb.active

            # 🔍 Tìm cột "Số chứng từ"
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            so_ct_idx = None
            for idx, col_name in enumerate(header_row):
                if str(col_name).strip().lower() == "số chứng từ":
                    so_ct_idx = idx
                    break

            if so_ct_idx is None:
                st.warning(f"⚠️ Không tìm thấy cột 'Số chứng từ' trong file: {filename}")
                continue

            # 🛠️ Duyệt từng dòng và chỉnh sửa nếu sai định dạng
            changes = 0
            for row in ws.iter_rows(min_row=2):
                cell = row[so_ct_idx]
                original_value = str(cell.value).strip() if cell.value else ""

                # Bỏ qua nếu đã đúng
                if original_value.startswith(prefix):
                    continue

                # Nếu là PTxxxxxx... → chuyển thành PT_THUOC_xxxxxx...
                if re.match(r"^PT\d", original_value):
                    new_value = prefix + original_value[2:]
                    cell.value = new_value
                    changes += 1

            # 💾 Lưu lại file Excel đã chỉnh sửa
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            st.success(f"✅ Đã xử lý file: {filename} ({changes} dòng được sửa)")
            st.download_button(
                label=f"⬇️ Tải xuống file đã sửa: {filename}",
                data=output,
                file_name=f"da_sua_{filename}",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"❌ Lỗi khi xử lý file {uploaded_file.name}: {e}")
